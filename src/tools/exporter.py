"""
Export utilities for generating PDF and Excel reports from AnalysisState data.

PDF:   fpdf2 (pure Python, no system dependencies)
Excel: openpyxl
"""

import io
import re
from datetime import datetime
from typing import Any


# ── Helpers ────────────────────────────────────────────────────────────────────

def _fmt_num(v: Any, decimals: int = 2) -> str:
    """Format a number or return 'N/A'."""
    if v is None:
        return "N/A"
    try:
        return f"{float(v):,.{decimals}f}"
    except (TypeError, ValueError):
        return str(v)


def _fmt_pct(v: Any) -> str:
    """Format a ratio (0.10 → 10.0%) or return 'N/A'."""
    if v is None:
        return "N/A"
    try:
        return f"{float(v) * 100:.1f}%"
    except (TypeError, ValueError):
        return str(v)


def _strip_markdown(text: str) -> str:
    """Remove basic markdown syntax and non-latin-1 chars for PDF output."""
    text = re.sub(r"#{1,6}\s*", "", text)
    text = re.sub(r"\*{1,3}(.+?)\*{1,3}", r"\1", text)
    text = re.sub(r"`(.+?)`", r"\1", text)
    text = re.sub(r"\[(.+?)\]\(.+?\)", r"\1", text)
    text = re.sub(r"^\s*[-*+]\s+", "- ", text, flags=re.MULTILINE)
    # Encode to latin-1 safe: replace unmappable chars with closest ASCII
    text = text.encode("latin-1", errors="replace").decode("latin-1")
    return text.strip()


def _safe_str(text: str) -> str:
    """Ensure a string is latin-1 encodable for fpdf2."""
    if not text:
        return ""
    return text.encode("latin-1", errors="replace").decode("latin-1")


# ── PDF Export ─────────────────────────────────────────────────────────────────

def generate_pdf(state: dict) -> bytes:
    """
    Generate a multi-section PDF investment report.
    Returns raw bytes suitable for an HTTP response.
    """
    from fpdf import FPDF

    ticker = state.get("ticker", "N/A")
    year = state.get("year", "")
    rating = (
        state.get("financial_statements", {}).get("investment_rating")
        or _extract_rating(state.get("final_report", ""))
        or "N/A"
    )
    generated_at = datetime.now().strftime("%Y-%m-%d %H:%M UTC")

    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.set_margins(15, 15, 15)

    # ── Cover Page ─────────────────────────────────────────────────────────────
    pdf.add_page()

    pdf.set_font("Helvetica", "B", 28)
    pdf.set_text_color(0, 188, 212)  # cyan
    pdf.ln(20)
    pdf.cell(0, 12, _safe_str(ticker), align="C", new_x="LMARGIN", new_y="NEXT")

    pdf.set_font("Helvetica", "", 14)
    pdf.set_text_color(180, 180, 180)
    pdf.cell(0, 8, f"Investment Analysis Report  ·  {year}", align="C", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(6)

    # Rating badge
    r_upper = rating.upper()
    if "BUY" in r_upper:
        pdf.set_fill_color(4, 120, 87)
    elif "SELL" in r_upper:
        pdf.set_fill_color(185, 28, 28)
    else:
        pdf.set_fill_color(120, 100, 0)
    pdf.set_text_color(255, 255, 255)
    pdf.set_font("Helvetica", "B", 16)
    pdf.cell(0, 10, f"  {rating}  ", align="C", fill=True, new_x="LMARGIN", new_y="NEXT")
    pdf.ln(4)

    # Live price summary
    live = state.get("live_price") or {}
    if live.get("price"):
        pdf.set_font("Helvetica", "", 11)
        pdf.set_text_color(200, 200, 200)
        price_line = f"Price: ${_fmt_num(live.get('price'))}  |  Change: {_fmt_num(live.get('change_pct'))}%  |  Mkt Cap: ${_fmt_num(live.get('market_cap', 0) / 1e9, 1)}B"
        pdf.cell(0, 7, price_line, align="C", new_x="LMARGIN", new_y="NEXT")
        pdf.ln(2)

    pdf.set_font("Helvetica", "", 9)
    pdf.set_text_color(120, 120, 120)
    pdf.cell(0, 6, f"Generated: {generated_at}", align="C", new_x="LMARGIN", new_y="NEXT")

    # Separator line
    pdf.ln(10)
    pdf.set_draw_color(0, 188, 212)
    pdf.set_line_width(0.5)
    pdf.line(15, pdf.get_y(), 195, pdf.get_y())

    # ── Final Report ───────────────────────────────────────────────────────────
    final_report = state.get("final_report", "")
    if final_report:
        pdf.add_page()
        _pdf_section_header(pdf, "Investment Report")
        _pdf_body_text(pdf, _strip_markdown(final_report))

    # ── Financial Statements ───────────────────────────────────────────────────
    stmts = state.get("financial_statements") or {}
    _pdf_financial_statements(pdf, stmts, ticker)

    # ── Valuation ─────────────────────────────────────────────────────────────
    val_data = state.get("valuation_data") or {}
    val_analysis = state.get("valuation_analysis", "")
    if val_data or val_analysis:
        pdf.add_page()
        _pdf_section_header(pdf, "Valuation Analysis")
        _pdf_valuation(pdf, val_data)
        if val_analysis:
            pdf.ln(4)
            _pdf_body_text(pdf, _strip_markdown(val_analysis))

    # ── Peer Comparison ────────────────────────────────────────────────────────
    peer_data = state.get("peer_data") or {}
    peer_analysis = state.get("peer_analysis", "")
    if peer_data or peer_analysis:
        pdf.add_page()
        _pdf_section_header(pdf, "Peer Comparison")
        _pdf_peers(pdf, peer_data)
        if peer_analysis:
            pdf.ln(4)
            _pdf_body_text(pdf, _strip_markdown(peer_analysis))

    # ── Debate Transcript ──────────────────────────────────────────────────────
    transcript = state.get("debate_transcript") or []
    if transcript:
        pdf.add_page()
        _pdf_section_header(pdf, "Analyst ↔ Critic Debate")
        _pdf_debate(pdf, transcript)

    return bytes(pdf.output())


def _pdf_section_header(pdf: "FPDF", title: str) -> None:
    pdf.set_font("Helvetica", "B", 14)
    pdf.set_text_color(0, 188, 212)
    pdf.cell(0, 8, _safe_str(title), new_x="LMARGIN", new_y="NEXT")
    pdf.set_draw_color(0, 188, 212)
    pdf.set_line_width(0.3)
    pdf.line(15, pdf.get_y(), 195, pdf.get_y())
    pdf.ln(4)
    pdf.set_text_color(30, 30, 30)


def _pdf_body_text(pdf: "FPDF", text: str) -> None:
    pdf.set_font("Helvetica", "", 9)
    pdf.set_text_color(40, 40, 40)
    for para in text.split("\n\n"):
        para = para.strip()
        if not para:
            continue
        pdf.set_x(pdf.l_margin)
        pdf.multi_cell(180, 5, para)
        pdf.ln(2)


def _pdf_table_header(pdf: "FPDF", cols: list[tuple[str, float]]) -> None:
    """cols: list of (label, width) tuples."""
    pdf.set_font("Helvetica", "B", 8)
    pdf.set_fill_color(30, 40, 55)
    pdf.set_text_color(0, 188, 212)
    for label, w in cols:
        pdf.cell(w, 6, _safe_str(label), border=0, fill=True, align="C")
    pdf.ln()
    pdf.set_text_color(40, 40, 40)


def _pdf_table_row(pdf: "FPDF", values: list[tuple[str, float]], alt: bool = False) -> None:
    pdf.set_font("Helvetica", "", 8)
    if alt:
        pdf.set_fill_color(245, 245, 250)
    else:
        pdf.set_fill_color(255, 255, 255)
    pdf.set_text_color(40, 40, 40)
    for val, w in values:
        pdf.cell(w, 5.5, _safe_str(str(val)), border=0, fill=True)
    pdf.ln()


def _pdf_financial_statements(pdf: "FPDF", stmts: dict, ticker: str) -> None:
    if not stmts:
        return

    income = stmts.get("income_statement") or {}
    balance = stmts.get("balance_sheet") or {}
    cashflow = stmts.get("cash_flow") or {}

    if not any([income, balance, cashflow]):
        return

    pdf.add_page()
    _pdf_section_header(pdf, "Financial Statements")

    for section_name, data in [
        ("Income Statement", income),
        ("Balance Sheet", balance),
        ("Cash Flow", cashflow),
    ]:
        if not data or not isinstance(data, dict):
            continue

        pdf.set_font("Helvetica", "B", 10)
        pdf.set_text_color(80, 80, 80)
        pdf.cell(0, 6, section_name, new_x="LMARGIN", new_y="NEXT")
        pdf.ln(1)

        cols = [("Metric", 100), ("Value", 80)]
        _pdf_table_header(pdf, cols)

        for i, (k, v) in enumerate(data.items()):
            if k in ("investment_rating", "fundamental_score"):
                continue
            label = k.replace("_", " ").title()
            val_str = _fmt_num(v) if isinstance(v, (int, float)) else str(v)
            _pdf_table_row(pdf, [(label, 100), (val_str, 80)], alt=i % 2 == 1)

        pdf.ln(4)


def _pdf_valuation(pdf: "FPDF", val_data: dict) -> None:
    dcf = val_data.get("dcf") or {}
    multiples = val_data.get("multiples") or {}
    overall = val_data.get("overall_verdict") or val_data.get("overall_assessment", "")

    if overall:
        pdf.set_font("Helvetica", "B", 11)
        pdf.set_text_color(0, 100, 150)
        pdf.cell(0, 7, _safe_str(f"Overall Verdict: {overall}"), new_x="LMARGIN", new_y="NEXT")
        pdf.ln(2)

    # DCF Table
    if dcf:
        pdf.set_font("Helvetica", "B", 10)
        pdf.set_text_color(80, 80, 80)
        pdf.cell(0, 6, "DCF Intrinsic Value", new_x="LMARGIN", new_y="NEXT")
        pdf.ln(1)
        dcf_rows = [
            ("Intrinsic Value ($/share)", f"${_fmt_num(dcf.get('intrinsic_value'))}"),
            ("Current Price ($/share)", f"${_fmt_num(dcf.get('current_price'))}"),
            ("Upside / Downside", f"{_fmt_num(dcf.get('upside_downside_pct'))}%"),
            ("WACC", _fmt_pct(dcf.get("wacc_used"))),
            ("Terminal Growth Rate", _fmt_pct(dcf.get("terminal_growth_rate"))),
        ]
        cols = [("Metric", 100), ("Value", 80)]
        _pdf_table_header(pdf, cols)
        for i, (k, v) in enumerate(dcf_rows):
            _pdf_table_row(pdf, [(k, 100), (v, 80)], alt=i % 2 == 1)
        pdf.ln(4)

        # Methodology
        if dcf.get("methodology"):
            pdf.set_font("Helvetica", "I", 8)
            pdf.set_text_color(100, 100, 100)
            pdf.set_x(pdf.l_margin)
            pdf.multi_cell(180, 4.5, _safe_str(f"Methodology: {dcf['methodology']}"))
            pdf.ln(3)

    # Multiples Table
    if multiples:
        pdf.set_font("Helvetica", "B", 10)
        pdf.set_text_color(80, 80, 80)
        pdf.cell(0, 6, "Relative Valuation Multiples", new_x="LMARGIN", new_y="NEXT")
        pdf.ln(1)

        MULTIPLE_LABELS = {
            "pe_trailing": "P/E (Trailing)",
            "pe_forward": "P/E (Forward)",
            "ev_ebitda": "EV/EBITDA",
            "price_to_sales": "P/S",
            "price_to_fcf": "P/FCF",
            "price_to_book": "P/B",
        }
        cols = [("Multiple", 65), ("Value", 45), ("Sector Avg", 45), ("Assessment", 35)]
        _pdf_table_header(pdf, cols)
        for i, (key, label) in enumerate(MULTIPLE_LABELS.items()):
            m = multiples.get(key) or {}
            if not m:
                continue
            row = [
                (label, 65),
                (_fmt_num(m.get("value")), 45),
                (_fmt_num(m.get("sector_avg")), 45),
                (str(m.get("assessment", "N/A")), 35),
            ]
            _pdf_table_row(pdf, row, alt=i % 2 == 1)
        pdf.ln(4)


def _pdf_peers(pdf: "FPDF", peer_data: dict) -> None:
    target = peer_data.get("target_company") or {}
    peers = peer_data.get("peers") or []
    positioning = peer_data.get("peer_analysis") or {}

    if not target and not peers:
        return

    # Positioning badge
    pos = positioning.get("overall_position", "")
    if pos:
        pdf.set_font("Helvetica", "B", 10)
        pdf.set_text_color(0, 100, 150)
        pdf.cell(0, 6, _safe_str(f"Competitive Position: {pos}"), new_x="LMARGIN", new_y="NEXT")
        pdf.ln(2)

    # Metrics table
    all_companies = ([target] if target else []) + list(peers)
    if not all_companies:
        return

    METRIC_COLS = [
        ("Company", 50),
        ("P/E", 22),
        ("Fwd P/E", 22),
        ("Margin%", 22),
        ("Rev Gr%", 22),
        ("D/E", 18),
        ("ROE%", 22),
    ]
    _pdf_table_header(pdf, METRIC_COLS)

    for i, co in enumerate(all_companies):
        name = (co.get("name") or co.get("ticker", ""))[:22]
        row = [
            (name, 50),
            (_fmt_num(co.get("pe_trailing"), 1), 22),
            (_fmt_num(co.get("pe_forward"), 1), 22),
            (_fmt_num(co.get("net_margin_pct"), 1), 22),
            (_fmt_num(co.get("revenue_growth_yoy_pct"), 1), 22),
            (_fmt_num(co.get("debt_to_equity"), 1), 18),
            (_fmt_num(co.get("roe_pct"), 1), 22),
        ]
        _pdf_table_row(pdf, row, alt=i % 2 == 1)

    pdf.ln(4)

    # Strengths / Weaknesses
    strengths = positioning.get("strengths") or []
    weaknesses = positioning.get("weaknesses") or []
    if strengths:
        pdf.set_font("Helvetica", "B", 9)
        pdf.set_text_color(4, 120, 87)
        pdf.cell(0, 5, "Strengths:", new_x="LMARGIN", new_y="NEXT")
        pdf.set_font("Helvetica", "", 8)
        pdf.set_text_color(40, 40, 40)
        for s in strengths:
            pdf.set_x(pdf.l_margin)
            pdf.multi_cell(180, 4.5, _safe_str(f"  - {s}"))
        pdf.ln(2)

    if weaknesses:
        pdf.set_font("Helvetica", "B", 9)
        pdf.set_text_color(185, 28, 28)
        pdf.cell(0, 5, "Weaknesses:", new_x="LMARGIN", new_y="NEXT")
        pdf.set_font("Helvetica", "", 8)
        pdf.set_text_color(40, 40, 40)
        for w in weaknesses:
            pdf.set_x(pdf.l_margin)
            pdf.multi_cell(180, 4.5, _safe_str(f"  - {w}"))
        pdf.ln(2)


def _pdf_debate(pdf: "FPDF", transcript: list) -> None:
    SPEAKER_COLORS = {
        "analyst": (0, 150, 100),
        "critic": (200, 80, 20),
    }
    for msg in transcript:
        speaker = msg.get("speaker", "")
        content = msg.get("content", "")
        rnd = msg.get("round", "")
        color = SPEAKER_COLORS.get(speaker, (60, 60, 60))

        pdf.set_font("Helvetica", "B", 9)
        pdf.set_text_color(*color)
        label = f"[Round {rnd}] {speaker.upper()}"
        pdf.cell(0, 5, label, new_x="LMARGIN", new_y="NEXT")

        pdf.set_font("Helvetica", "", 8)
        pdf.set_text_color(40, 40, 40)
        pdf.set_x(pdf.l_margin)
        pdf.multi_cell(180, 4.5, _strip_markdown(content))
        pdf.ln(3)


def _extract_rating(text: str) -> str:
    m = re.search(r"\b(STRONG BUY|STRONG SELL|BUY|SELL|HOLD)\b", text, re.IGNORECASE)
    return m.group(1).upper() if m else ""


# ── Excel Export ───────────────────────────────────────────────────────────────

def generate_excel(state: dict) -> bytes:
    """
    Generate a multi-sheet Excel workbook.
    Returns raw bytes suitable for an HTTP response.
    """
    from openpyxl import Workbook
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, numbers
    )

    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    ticker = state.get("ticker", "N/A")

    # Helper styles
    HDR_FILL = PatternFill("solid", fgColor="0D1B2A")
    HDR_FONT = Font(bold=True, color="00BCD4", size=10)
    SUBHDR_FILL = PatternFill("solid", fgColor="1E2A3A")
    SUBHDR_FONT = Font(bold=True, color="90CAF9", size=9)
    ALT_FILL = PatternFill("solid", fgColor="F5F5FA")
    TITLE_FONT = Font(bold=True, color="00BCD4", size=14)
    BOLD = Font(bold=True, size=9)
    NORMAL = Font(size=9)
    CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="CCCCCC")
    BORDER = Border(bottom=thin)

    def style_header_row(ws, row: int, cols: int) -> None:
        for c in range(1, cols + 1):
            cell = ws.cell(row=row, column=c)
            cell.fill = HDR_FILL
            cell.font = HDR_FONT
            cell.alignment = CENTER
            cell.border = BORDER

    def style_subheader_row(ws, row: int, cols: int) -> None:
        for c in range(1, cols + 1):
            cell = ws.cell(row=row, column=c)
            cell.fill = SUBHDR_FILL
            cell.font = SUBHDR_FONT
            cell.alignment = LEFT
            cell.border = BORDER

    def style_data_row(ws, row: int, cols: int, alt: bool = False) -> None:
        for c in range(1, cols + 1):
            cell = ws.cell(row=row, column=c)
            if alt:
                cell.fill = ALT_FILL
            cell.font = NORMAL
            cell.alignment = LEFT

    # ── Sheet 1: Summary ───────────────────────────────────────────────────────
    ws = wb.create_sheet("Summary")
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 30

    ws.merge_cells("A1:B1")
    ws["A1"] = f"{ticker} — Investment Analysis Summary"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = CENTER
    ws.row_dimensions[1].height = 22

    stmts = state.get("financial_statements") or {}
    live = state.get("live_price") or {}
    rating = stmts.get("investment_rating") or _extract_rating(state.get("final_report", ""))

    summary_rows = [
        ("Ticker", ticker),
        ("Analysis Year", str(state.get("year", ""))),
        ("Investment Rating", rating or "N/A"),
        ("Fundamental Score", stmts.get("fundamental_score", "N/A")),
        ("Current Price", f"${_fmt_num(live.get('price'))}" if live.get("price") else "N/A"),
        ("Price Change %", f"{_fmt_num(live.get('change_pct'))}%" if live.get("change_pct") is not None else "N/A"),
        ("52W High", f"${_fmt_num(live.get('week52_high'))}" if live.get("week52_high") else "N/A"),
        ("52W Low", f"${_fmt_num(live.get('week52_low'))}" if live.get("week52_low") else "N/A"),
        ("Market Cap", f"${_fmt_num((live.get('market_cap') or 0) / 1e9, 1)}B" if live.get("market_cap") else "N/A"),
        ("Generated At", datetime.now().strftime("%Y-%m-%d %H:%M UTC")),
    ]

    ws.cell(row=2, column=1, value="Field").font = SUBHDR_FONT
    ws.cell(row=2, column=1).fill = SUBHDR_FILL
    ws.cell(row=2, column=2, value="Value").font = SUBHDR_FONT
    ws.cell(row=2, column=2).fill = SUBHDR_FILL

    for i, (k, v) in enumerate(summary_rows):
        r = i + 3
        ws.cell(row=r, column=1, value=k).font = BOLD
        ws.cell(row=r, column=2, value=v).font = NORMAL
        if i % 2 == 1:
            ws.cell(row=r, column=1).fill = ALT_FILL
            ws.cell(row=r, column=2).fill = ALT_FILL

    # Final report text below
    final_report = state.get("final_report", "")
    if final_report:
        r_start = len(summary_rows) + 4
        ws.merge_cells(f"A{r_start}:B{r_start}")
        ws.cell(row=r_start, column=1, value="Final Investment Report").font = SUBHDR_FONT
        ws.cell(row=r_start, column=1).fill = SUBHDR_FILL
        ws.cell(row=r_start + 1, column=1, value=_strip_markdown(final_report)).font = NORMAL
        ws.cell(row=r_start + 1, column=1).alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(f"A{r_start+1}:B{r_start+1}")
        ws.row_dimensions[r_start + 1].height = max(40, min(len(final_report) // 4, 300))

    # ── Sheet 2: Financial Statements ──────────────────────────────────────────
    ws_fin = wb.create_sheet("Financial Statements")
    ws_fin.column_dimensions["A"].width = 35
    ws_fin.column_dimensions["B"].width = 22

    ws_fin.merge_cells("A1:B1")
    ws_fin["A1"] = f"{ticker} — Financial Statements"
    ws_fin["A1"].font = TITLE_FONT
    ws_fin["A1"].alignment = CENTER
    ws_fin.row_dimensions[1].height = 20

    row_cur = 2
    for section_name, data in [
        ("Income Statement", stmts.get("income_statement") or {}),
        ("Balance Sheet", stmts.get("balance_sheet") or {}),
        ("Cash Flow", stmts.get("cash_flow") or {}),
    ]:
        if not data or not isinstance(data, dict):
            continue

        ws_fin.merge_cells(f"A{row_cur}:B{row_cur}")
        ws_fin.cell(row=row_cur, column=1, value=section_name)
        style_subheader_row(ws_fin, row_cur, 2)
        row_cur += 1

        ws_fin.cell(row=row_cur, column=1, value="Metric").font = BOLD
        ws_fin.cell(row=row_cur, column=2, value="Value").font = BOLD
        style_header_row(ws_fin, row_cur, 2)
        row_cur += 1

        for i, (k, v) in enumerate(data.items()):
            if k in ("investment_rating", "fundamental_score"):
                continue
            ws_fin.cell(row=row_cur, column=1, value=k.replace("_", " ").title())
            ws_fin.cell(row=row_cur, column=2, value=v if isinstance(v, (int, float)) else str(v))
            style_data_row(ws_fin, row_cur, 2, alt=i % 2 == 1)
            row_cur += 1

        row_cur += 1

    # ── Sheet 3: Valuation ─────────────────────────────────────────────────────
    ws_val = wb.create_sheet("Valuation")
    ws_val.column_dimensions["A"].width = 30
    ws_val.column_dimensions["B"].width = 20
    ws_val.column_dimensions["C"].width = 20
    ws_val.column_dimensions["D"].width = 18

    ws_val.merge_cells("A1:D1")
    ws_val["A1"] = f"{ticker} — Valuation Analysis"
    ws_val["A1"].font = TITLE_FONT
    ws_val["A1"].alignment = CENTER
    ws_val.row_dimensions[1].height = 20

    val_data = state.get("valuation_data") or {}
    dcf = val_data.get("dcf") or {}
    multiples = val_data.get("multiples") or {}
    overall = val_data.get("overall_verdict") or val_data.get("overall_assessment", "")

    row_cur = 2
    if overall:
        ws_val.merge_cells(f"A{row_cur}:D{row_cur}")
        ws_val.cell(row=row_cur, column=1, value=f"Overall Verdict: {overall}")
        ws_val.cell(row=row_cur, column=1).font = Font(bold=True, color="00BCD4", size=11)
        ws_val.cell(row=row_cur, column=1).alignment = CENTER
        row_cur += 1

    # DCF section
    if dcf:
        ws_val.merge_cells(f"A{row_cur}:D{row_cur}")
        ws_val.cell(row=row_cur, column=1, value="DCF Intrinsic Value")
        style_subheader_row(ws_val, row_cur, 4)
        row_cur += 1

        dcf_rows = [
            ("Intrinsic Value ($/share)", f"${_fmt_num(dcf.get('intrinsic_value'))}"),
            ("Current Price ($/share)", f"${_fmt_num(dcf.get('current_price'))}"),
            ("Upside / Downside", f"{_fmt_num(dcf.get('upside_downside_pct'))}%"),
            ("WACC", _fmt_pct(dcf.get("wacc_used"))),
            ("Terminal Growth Rate", _fmt_pct(dcf.get("terminal_growth_rate"))),
            ("Terminal Value ($B)", f"${_fmt_num(dcf.get('terminal_value'))}B"),
        ]
        for i, (k, v) in enumerate(dcf_rows):
            ws_val.cell(row=row_cur, column=1, value=k).font = BOLD
            ws_val.cell(row=row_cur, column=2, value=v).font = NORMAL
            style_data_row(ws_val, row_cur, 4, alt=i % 2 == 1)
            row_cur += 1

        if dcf.get("methodology"):
            ws_val.merge_cells(f"A{row_cur}:D{row_cur}")
            ws_val.cell(row=row_cur, column=1, value=f"Methodology: {dcf['methodology']}")
            ws_val.cell(row=row_cur, column=1).font = Font(italic=True, size=8, color="666666")
            ws_val.cell(row=row_cur, column=1).alignment = Alignment(wrap_text=True)
            row_cur += 1

        row_cur += 1

    # Multiples section
    if multiples:
        ws_val.merge_cells(f"A{row_cur}:D{row_cur}")
        ws_val.cell(row=row_cur, column=1, value="Relative Valuation Multiples")
        style_subheader_row(ws_val, row_cur, 4)
        row_cur += 1

        for col, label in enumerate(["Multiple", "Value", "Sector Avg", "Assessment"], 1):
            ws_val.cell(row=row_cur, column=col, value=label)
        style_header_row(ws_val, row_cur, 4)
        row_cur += 1

        MULTIPLE_LABELS = {
            "pe_trailing": "P/E (Trailing)",
            "pe_forward": "P/E (Forward)",
            "ev_ebitda": "EV/EBITDA",
            "price_to_sales": "P/S",
            "price_to_fcf": "P/FCF",
            "price_to_book": "P/B",
        }
        for i, (key, label) in enumerate(MULTIPLE_LABELS.items()):
            m = multiples.get(key) or {}
            if not m:
                continue
            ws_val.cell(row=row_cur, column=1, value=label)
            ws_val.cell(row=row_cur, column=2, value=_fmt_num(m.get("value")))
            ws_val.cell(row=row_cur, column=3, value=_fmt_num(m.get("sector_avg")))
            ws_val.cell(row=row_cur, column=4, value=str(m.get("assessment", "N/A")))
            style_data_row(ws_val, row_cur, 4, alt=i % 2 == 1)
            row_cur += 1

    # ── Sheet 4: Peer Comparison ───────────────────────────────────────────────
    ws_peer = wb.create_sheet("Peer Comparison")
    peer_data = state.get("peer_data") or {}
    target_co = peer_data.get("target_company") or {}
    peers = peer_data.get("peers") or []
    positioning = peer_data.get("peer_analysis") or {}

    PEER_HEADERS = ["Ticker", "Company", "P/E", "Fwd P/E", "Net Margin%",
                    "Rev Growth%", "D/E Ratio", "ROE%", "Mkt Cap ($B)", "Analyst Target"]
    PEER_WIDTHS = [12, 25, 10, 10, 14, 14, 12, 10, 14, 16]

    for col_i, (h, w) in enumerate(zip(PEER_HEADERS, PEER_WIDTHS), 1):
        ws_peer.column_dimensions[ws_peer.cell(1, col_i).column_letter].width = w

    ws_peer.merge_cells(f"A1:{ws_peer.cell(1, len(PEER_HEADERS)).column_letter}1")
    ws_peer["A1"] = f"{ticker} — Peer Comparison"
    ws_peer["A1"].font = TITLE_FONT
    ws_peer["A1"].alignment = CENTER
    ws_peer.row_dimensions[1].height = 20

    for col_i, h in enumerate(PEER_HEADERS, 1):
        ws_peer.cell(row=2, column=col_i, value=h)
    style_header_row(ws_peer, 2, len(PEER_HEADERS))

    def _peer_row(ws, row, co, alt):
        mktcap = co.get("market_cap_billions")
        ws.cell(row=row, column=1, value=co.get("ticker", ""))
        ws.cell(row=row, column=2, value=co.get("name", ""))
        ws.cell(row=row, column=3, value=_fmt_num(co.get("pe_trailing"), 1))
        ws.cell(row=row, column=4, value=_fmt_num(co.get("pe_forward"), 1))
        ws.cell(row=row, column=5, value=_fmt_num(co.get("net_margin_pct"), 1))
        ws.cell(row=row, column=6, value=_fmt_num(co.get("revenue_growth_yoy_pct"), 1))
        ws.cell(row=row, column=7, value=_fmt_num(co.get("debt_to_equity"), 2))
        ws.cell(row=row, column=8, value=_fmt_num(co.get("roe_pct"), 1))
        ws.cell(row=row, column=9, value=_fmt_num(mktcap, 1) if mktcap is not None else "N/A")
        ws.cell(row=row, column=10, value=_fmt_num(co.get("analyst_target"), 2))
        style_data_row(ws, row, len(PEER_HEADERS), alt=alt)

    row_cur = 3
    if target_co:
        _peer_row(ws_peer, row_cur, target_co, alt=False)
        # bold the target row
        for c in range(1, len(PEER_HEADERS) + 1):
            ws_peer.cell(row=row_cur, column=c).font = Font(bold=True, size=9)
        row_cur += 1

    for i, peer in enumerate(peers):
        _peer_row(ws_peer, row_cur, peer, alt=i % 2 == 0)
        row_cur += 1

    # Competitive positioning
    if positioning:
        row_cur += 1
        pos_label = positioning.get("overall_position", "")
        if pos_label:
            ws_peer.cell(row=row_cur, column=1, value=f"Competitive Position: {pos_label}")
            ws_peer.cell(row=row_cur, column=1).font = Font(bold=True, color="00BCD4", size=11)
            row_cur += 1

        for label, items_key in [("Strengths", "strengths"), ("Weaknesses", "weaknesses")]:
            items = positioning.get(items_key) or []
            if items:
                ws_peer.cell(row=row_cur, column=1, value=label).font = BOLD
                row_cur += 1
                for item in items:
                    ws_peer.cell(row=row_cur, column=1, value=f"  • {item}").font = NORMAL
                    ws_peer.merge_cells(f"A{row_cur}:{ws_peer.cell(row_cur, len(PEER_HEADERS)).column_letter}{row_cur}")
                    ws_peer.cell(row=row_cur, column=1).alignment = Alignment(wrap_text=True)
                    row_cur += 1

    # ── Sheet 5: Debate Transcript ─────────────────────────────────────────────
    transcript = state.get("debate_transcript") or []
    if transcript:
        ws_dbt = wb.create_sheet("Debate Transcript")
        ws_dbt.column_dimensions["A"].width = 12
        ws_dbt.column_dimensions["B"].width = 15
        ws_dbt.column_dimensions["C"].width = 90

        ws_dbt.merge_cells("A1:C1")
        ws_dbt["A1"] = f"{ticker} — Analyst ↔ Critic Debate"
        ws_dbt["A1"].font = TITLE_FONT
        ws_dbt["A1"].alignment = CENTER
        ws_dbt.row_dimensions[1].height = 20

        for col_i, h in enumerate(["Round", "Speaker", "Content"], 1):
            ws_dbt.cell(row=2, column=col_i, value=h)
        style_header_row(ws_dbt, 2, 3)

        SPEAKER_COLORS_XLSX = {"analyst": "04786C", "critic": "C85014"}
        for i, msg in enumerate(transcript):
            r = i + 3
            ws_dbt.cell(row=r, column=1, value=msg.get("round", ""))
            speaker = msg.get("speaker", "")
            ws_dbt.cell(row=r, column=2, value=speaker.upper())
            ws_dbt.cell(row=r, column=2).font = Font(
                bold=True, size=9, color=SPEAKER_COLORS_XLSX.get(speaker, "333333")
            )
            ws_dbt.cell(row=r, column=3, value=_strip_markdown(msg.get("content", "")))
            ws_dbt.cell(row=r, column=3).alignment = Alignment(wrap_text=True, vertical="top")
            ws_dbt.row_dimensions[r].height = 60
            if i % 2 == 1:
                ws_dbt.cell(row=r, column=1).fill = ALT_FILL
                ws_dbt.cell(row=r, column=3).fill = ALT_FILL

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
